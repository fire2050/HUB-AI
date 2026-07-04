from __future__ import annotations

import re
import sqlite3
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.db import DB_PATH  # noqa: E402

try:
    import openpyxl
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl is required: {exc}")

EXCEL_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("wuying_products.xlsx")
DOMESTIC_COMPUTE_SHEETS = [
    "D21国内120小时",
    "D21国内不限时长-办公1小时休眠",
    "D22国内图形200小时",
    "D22国内不限时长(1小时休眠)",
    "D23国内不限时长",
]


def norm(v: Any) -> str:
    return str(v or "").strip()


def to_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("￥", "").replace("元", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else default


def slug(text: str) -> str:
    s = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff]+", "-", text).strip("-")
    return s[:90]


def find_domestic_compute_sections(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    sections: list[dict] = []
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        joined = " ".join(vals)
        if "国内region计算规格名称" in joined:
            header_row = idx
            period_row = idx + 1
            sections.append({"header_row": header_row, "period_row": period_row})
    return sections


# Sheet → product_type / category mapping
SHEET_MAP: list[dict] = [
    # 国内计算规格
    {"name": "D21国内120小时", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "domestic", "category_code": "CAT_WUYING_DOMESTIC_COMPUTE", "category_name": "无影国内计算规格"},
    {"name": "D21国内不限时长-办公1小时休眠", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "domestic", "category_code": "CAT_WUYING_DOMESTIC_COMPUTE", "category_name": "无影国内计算规格"},
    {"name": "D22国内图形200小时", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "domestic", "category_code": "CAT_WUYING_DOMESTIC_COMPUTE", "category_name": "无影国内计算规格"},
    {"name": "D22国内不限时长(1小时休眠)", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "domestic", "category_code": "CAT_WUYING_DOMESTIC_COMPUTE", "category_name": "无影国内计算规格"},
    {"name": "D23国内不限时长", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "domestic", "category_code": "CAT_WUYING_DOMESTIC_COMPUTE", "category_name": "无影国内计算规格"},
    # 国际计算规格
    {"name": "D31国际120小时", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "international", "category_code": "CAT_WUYING_INTL_COMPUTE", "category_name": "无影国际计算规格"},
    {"name": "D32国际200小时", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "international", "category_code": "CAT_WUYING_INTL_COMPUTE", "category_name": "无影国际计算规格"},
    {"name": "D33国际不限时长", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "international", "category_code": "CAT_WUYING_INTL_COMPUTE", "category_name": "无影国际计算规格"},
    # 教育办公
    {"name": "D1教育办公", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "domestic", "category_code": "CAT_WUYING_EDU_COMPUTE", "category_name": "无影教育办公计算规格"},
    # 企业版系统规格
    {"name": "B1企业版系统规格list价格（必选项）", "product_type": "compute", "product_type_name": "计算规格", "region_scope": "enterprise", "category_code": "CAT_WUYING_ENT_COMPUTE", "category_name": "无影企业版计算规格"},
    # 磁盘
    {"name": "B2-磁盘规格list价格（必选项）", "product_type": "storage", "product_type_name": "磁盘规格", "region_scope": "common", "category_code": "CAT_WUYING_STORAGE", "category_name": "无影磁盘规格"},
    # 带宽
    {"name": "B3-带宽规格list价格（可选项）", "product_type": "bandwidth", "product_type_name": "带宽规格", "region_scope": "common", "category_code": "CAT_WUYING_BANDWIDTH", "category_name": "无影带宽规格"},
    {"name": "D2-带宽规格list价格（可选项） ", "product_type": "bandwidth", "product_type_name": "带宽规格", "region_scope": "common", "category_code": "CAT_WUYING_BANDWIDTH", "category_name": "无影带宽规格"},
    # 流量包
    {"name": "B4-流量包list价格（可选项）", "product_type": "traffic_package", "product_type_name": "流量包", "region_scope": "common", "category_code": "CAT_WUYING_TRAFFIC", "category_name": "无影流量包"},
    # AD Connector
    {"name": "B5-AD Connector的list价格（可选项）", "product_type": "addon", "product_type_name": "AD Connector", "region_scope": "common", "category_code": "CAT_WUYING_ADDON", "category_name": "无影AD Connector"},
    # 企业网盘
    {"name": "B6-企业网盘list价格（可选项）", "product_type": "addon", "product_type_name": "企业网盘", "region_scope": "common", "category_code": "CAT_WUYING_CLOUD_DRIVE", "category_name": "无影企业网盘"},
    # 核时包
    {"name": "B7-核时包list价格（可选项）", "product_type": "core_hour_package", "product_type_name": "核时包", "region_scope": "common", "category_code": "CAT_WUYING_CORE_HOUR", "category_name": "无影核时包"},
    # 终端
    {"name": "C0-终端描述汇总", "product_type": "terminal", "product_type_name": "终端外设", "region_scope": "common", "category_code": "CAT_WUYING_TERMINAL", "category_name": "无影终端外设"},
    # AI 助手
    {"name": "E贾维斯", "product_type": "ai_assistant", "product_type_name": "AI助手", "region_scope": "common", "category_code": "CAT_WUYING_AI", "category_name": "无影AI助手"},
]


def import_wuying_excel(excel_path: Path) -> dict:
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 清空产品报价域数据。先删除 FTS 触发器，避免旧 FTS 内容与批量删除冲突。
    for trigger in ["knowledge_chunk_ai", "knowledge_chunk_ad", "knowledge_chunk_au"]:
        try:
            cur.execute(f"DROP TRIGGER IF EXISTS {trigger}")
        except sqlite3.OperationalError:
            pass

    for table in [
        "knowledge_chunk_fts", "knowledge_chunk", "product_faq", "product_spec", "product_document",
        "inventory", "product_price", "quotation_line", "quotation_header", "quotation_policy", "product", "product_category",
    ]:
        try:
            cur.execute(f"DELETE FROM {table}")
        except sqlite3.OperationalError:
            pass

    # 默认不打折
    cur.execute(
        "INSERT OR REPLACE INTO quotation_policy (policy_code, policy_name, discount_rate, min_quantity, min_amount, status) VALUES (?, ?, ?, ?, ?, ?)",
        ("POL-WUYING-NO-DISCOUNT", "无影默认无折扣报价政策", 1.0, 1, 0, "active"),
    )

    imported_products = 0
    imported_prices = 0
    imported_specs = 0
    skipped_rows = 0
    sheet_stats: list[dict] = []

    for cfg in SHEET_MAP:
        name = cfg["name"]
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        cur.execute(
            "INSERT OR IGNORE INTO product_category (category_code, category_name, description) VALUES (?, ?, ?)",
            (cfg["category_code"], cfg["category_name"], f"由无影 Excel sheet {name} 导入"),
        )

        # Create a sheet-specific importer (special handling per sheet type)
        if cfg["product_type"] == "compute":
            stats = _import_compute_sheet(wb[name], cfg, cur)
        elif cfg["product_type"] == "storage":
            stats = _import_storage_sheet(wb[name], cfg, cur)
        elif cfg["product_type"] == "bandwidth":
            stats = _import_bandwidth_sheet(wb[name], cfg, cur)
        elif cfg["product_type"] == "traffic_package":
            stats = _import_traffic_sheet(wb[name], cfg, cur)
        elif cfg["product_type"] == "addon":
            stats = _import_addon_sheet(wb[name], cfg, cur)
        elif cfg["product_type"] == "core_hour_package":
            stats = _import_core_hour_sheet(wb[name], cfg, cur)
        elif cfg["product_type"] == "terminal":
            stats = _import_terminal_sheet(wb[name], cfg, cur)
        elif cfg["product_type"] == "ai_assistant":
            stats = _import_ai_sheet(wb[name], cfg, cur)
        else:
            stats = {"products": 0, "prices": 0, "specs": 0, "skipped": 0}

        imported_products += stats["products"]
        imported_prices += stats["prices"]
        imported_specs += stats["specs"]
        skipped_rows += stats["skipped"]
        sheet_stats.append({"sheet": name, "products": stats["products"]})

    # Recreate FTS triggers after bulk import so later knowledge_chunk changes stay searchable.
    try:
        cur.executescript(
            """
            CREATE TRIGGER IF NOT EXISTS knowledge_chunk_ai AFTER INSERT ON knowledge_chunk BEGIN
              INSERT INTO knowledge_chunk_fts(rowid, chunk_text, title, source_type, source_id)
              VALUES (NEW.id, NEW.chunk_text, NEW.title, NEW.source_type, NEW.source_id);
            END;
            CREATE TRIGGER IF NOT EXISTS knowledge_chunk_ad AFTER DELETE ON knowledge_chunk BEGIN
              INSERT INTO knowledge_chunk_fts(knowledge_chunk_fts, rowid, chunk_text, title, source_type, source_id)
              VALUES ('delete', OLD.id, OLD.chunk_text, OLD.title, OLD.source_type, OLD.source_id);
            END;
            CREATE TRIGGER IF NOT EXISTS knowledge_chunk_au AFTER UPDATE ON knowledge_chunk BEGIN
              INSERT INTO knowledge_chunk_fts(knowledge_chunk_fts, rowid, chunk_text, title, source_type, source_id)
              VALUES ('delete', OLD.id, OLD.chunk_text, OLD.title, OLD.source_type, OLD.source_id);
              INSERT INTO knowledge_chunk_fts(rowid, chunk_text, title, source_type, source_id)
              VALUES (NEW.id, NEW.chunk_text, NEW.title, NEW.source_type, NEW.source_id);
            END;
            """
        )
    except sqlite3.OperationalError:
        pass

    conn.commit()
    counts = {}
    for table in ["product", "product_price", "inventory", "product_spec", "knowledge_chunk", "quotation_policy", "product_category"]:
        counts[table] = cur.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
    price_types = [dict(r) for r in cur.execute("SELECT price_type, COUNT(*) AS count, MIN(unit_price) AS min_price, MAX(unit_price) AS max_price FROM product_price GROUP BY price_type ORDER BY price_type")]
    conn.close()
    return {
        "excel": str(excel_path),
        "imported_products": imported_products,
        "imported_prices": imported_prices,
        "imported_specs": imported_specs,
        "skipped_rows": skipped_rows,
        "sheet_stats": sheet_stats,
        "counts": counts,
        "price_types": price_types,
    }


# ── Sheet-specific importers ──────────────────────────────────────────

def _insert_product(
    cur,
    product_code: str,
    product_name: str,
    cfg: dict,
    config_desc: str,
    model: str = "",
    unit: str = "套",
    ext_json: dict | None = None,
) -> None:
    import json
    cur.execute(
        """
        INSERT OR REPLACE INTO product
        (product_code, product_name, category_code, product_type, product_type_name, region_scope,
         source_sheet, brand, model, unit, product_config_description, short_description, status, ext_json)
        VALUES (?, ?, ?, ?, ?, ?, ?, '无影', ?, ?, ?, ?, 'active', ?)
        """,
        (product_code, product_name, cfg["category_code"], cfg["product_type"], cfg["product_type_name"],
         cfg["region_scope"], cfg["name"], model, unit, config_desc, config_desc,
         json.dumps(ext_json or {}, ensure_ascii=False)),
    )


def _insert_price(
    cur, product_code: str, unit_price: float, price_type: str, billing_period: str,
    cfg: dict, source_column: str = "", unit_label: str = "", allow_zero: bool = False
) -> None:
    import json
    if unit_price > 0 or allow_zero:
        cur.execute(
            """
            INSERT INTO product_price
            (product_code, unit_price, currency, price_type, billing_period,
             source_sheet, source_column, unit_label, valid_from, status, ext_json)
            VALUES (?, ?, 'CNY', ?, ?, ?, ?, ?, NULL, 'active', ?)
            """,
            (product_code, unit_price, price_type, billing_period, cfg["name"], source_column,
             unit_label, json.dumps({"excel_column": source_column, "source_sheet": cfg["name"]}, ensure_ascii=False)),
        )


def _insert_inventory(cur, product_code: str) -> None:
    cur.execute(
        "INSERT INTO inventory (product_code, warehouse_code, quantity, reserved_quantity, safety_stock) VALUES (?, 'CN-DOMESTIC', 999, 0, 10)",
        (product_code,),
    )


def _insert_spec(cur, product_code: str, name: str, value: Any, unit: str | None, group: str, order: int) -> None:
    if value not in (None, ""):
        cur.execute(
            "INSERT INTO product_spec (product_code, spec_name, spec_value, spec_unit, spec_group, sort_order) VALUES (?, ?, ?, ?, ?, ?)",
            (product_code, name, str(value), unit, group, order),
        )


def _insert_chunk(cur, product_code: str, product_name: str, config_desc: str, tags: str) -> None:
    chunk = f"产品名称：{product_name}\n产品配置描述：{config_desc}"
    cur.execute(
        "INSERT INTO knowledge_chunk (source_type, source_id, title, chunk_text, chunk_index, tags) VALUES ('product_doc', ?, ?, ?, 0, ?)",
        (product_code, product_name, chunk, tags),
    )


# ── Compute sheet importer ────────────────────────────────────────────

def _import_compute_sheet(ws, cfg: dict, cur) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    sections = _find_compute_sections(rows)
    products = prices = specs = skipped = 0
    for sec in sections:
        name_col = sec.get("name_col", 0)
        header = [norm(x) for x in rows[sec["header_row"]]]
        periods = [norm(x) for x in rows[sec["period_row"]]] if sec["period_row"] < len(rows) else []
        desc_col = None
        remark_col = None
        for i, h in enumerate(header):
            if "描述" in h:
                desc_col = i
            if "备注" in h:
                remark_col = i
        # Find period columns: "一个月", "一年", "二年", "三年", "四年", "五年", "六年", "不限时", "360小时", "250小时", "120小时", "目录小时价"
        period_cols: list[tuple[int, str, str]] = []
        for i, p in enumerate(periods):
            p_n = p.replace("\n", "").replace("月", "月")
            if p_n == "一个月":
                period_cols.append((i, "1month", "一个月"))
            elif p_n == "一年":
                period_cols.append((i, "1year", "一年"))
            elif p_n == "二年":
                period_cols.append((i, "2year", "二年"))
            elif p_n == "三年":
                period_cols.append((i, "3year", "三年"))
            elif p_n == "四年":
                period_cols.append((i, "4year", "四年"))
            elif p_n == "五年":
                period_cols.append((i, "5year", "五年"))
            elif p_n == "六年":
                period_cols.append((i, "6year", "六年"))
            elif "不限时" in p_n:
                period_cols.append((i, "unlimited", "不限时"))
            elif "小时" in p_n and "目录" in p_n:
                period_cols.append((i, "hourly", "目录小时价"))
        # vCPU / mem / gpu / disk positions from header
        vcpu_col = _find_col(header, ["vCPU"])
        mem_col = _find_col(header, ["内存", "内存(GiB)"])
        gpu_col = _find_col(header, ["显存", "显存(GiB)"])
        disk_col = _find_col(header, ["磁盘", "磁盘(GiB)"])
        for r_idx, row in enumerate(rows[sec["period_row"] + 1:], start=sec["period_row"] + 1):
            vals = [norm(x) for x in row]
            if not any(vals):
                continue
            spec_name = vals[name_col] if name_col < len(vals) else ""
            if not spec_name:
                skipped += 1
                continue
            if any(k in spec_name for k in ["计算规格名称", "存储规格名称", "带宽规格名称"]):
                continue
            if any(k in spec_name for k in ["国内region存储规格", "国内region精品带宽规格", "国际region存储规格", "国际region带宽规格"]):
                break
            config_desc = vals[desc_col] if desc_col is not None and desc_col < len(vals) else ""
            remark = vals[remark_col] if remark_col is not None and remark_col < len(vals) else ""
            vcpu = vals[vcpu_col] if vcpu_col is not None and vcpu_col < len(vals) else ""
            mem = vals[mem_col] if mem_col is not None and mem_col < len(vals) else ""
            gpu = vals[gpu_col] if gpu_col is not None and gpu_col < len(vals) else ""
            disk = vals[disk_col] if disk_col is not None and disk_col < len(vals) else ""
            product_code = f"WY-{slug(cfg['category_code'])}-{slug(cfg['name'])}-{slug(spec_name)}"
            _insert_product(
                cur, product_code, spec_name, cfg, config_desc,
                model=f"{vcpu}核/{mem}GiB/显存{gpu}/磁盘{disk}",
                unit="台",
                ext_json={"vCPU": vcpu, "memory": mem, "gpu": gpu, "disk": disk, "remark": remark},
            )
            products += 1
            _insert_chunk(cur, product_code, spec_name, config_desc, f"无影,{cfg['product_type_name']},{cfg['region_scope']},计算规格")
            _insert_inventory(cur, product_code)
            for order, (sn, sv, su) in enumerate([
                ("vCPU", vcpu, "核"),
                ("内存", mem, "GiB"),
                ("显存", gpu, "GiB"),
                ("磁盘", disk, "GiB"),
                ("产品配置描述", config_desc, None),
            ], start=1):
                _insert_spec(cur, product_code, sn, sv, su, cfg["product_type_name"], order)
                specs += 1
            for col_i, price_type, col_label in period_cols:
                price_val = to_float(vals[col_i] if col_i < len(vals) else "")
                _insert_price(cur, product_code, price_val, price_type, col_label, cfg, col_label, "元/月/台")
                if price_val > 0:
                    prices += 1
    return {"products": products, "prices": prices, "specs": specs, "skipped": skipped}


def _find_compute_sections(rows: list[list[Any]]) -> list[dict]:
    sections: list[dict] = []
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        joined = " ".join(vals)
        if any(k in joined for k in ["计算规格名称", "计算规格（必选项）"]):
            name_col = 0
            for i, v in enumerate(vals):
                if "名称" in v:
                    name_col = i
                    break
            sections.append({"header_row": idx, "period_row": idx + 1, "name_col": name_col})
    return sections


def _find_col(header: list[str], keywords: list[str]) -> int | None:
    for i, h in enumerate(header):
        for kw in keywords:
            if kw in h:
                return i
    return None


# ── Storage (disk) importer ───────────────────────────────────────────

def _import_storage_sheet(ws, cfg: dict, cur) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    products = prices = specs = skipped = 0
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        if not any(vals):
            continue
        if "存储规格" in vals[0] or "磁盘规格" in vals[0]:
            continue
        name = vals[0]
        if not name:
            skipped += 1
            continue
        monthly_price = to_float(vals[1] if len(vals) > 1 else "")
        hourly_price = to_float(vals[2] if len(vals) > 2 else "")
        desc = vals[3] if len(vals) > 3 else ""
        product_code = f"WY-{slug(cfg['category_code'])}-{slug(name)}"
        _insert_product(cur, product_code, name, cfg, desc, model=name, unit="GiB", ext_json={"storage_type": name})
        products += 1
        _insert_inventory(cur, product_code)
        _insert_chunk(cur, product_code, name, desc, "无影,磁盘规格")
        _insert_price(cur, product_code, monthly_price, "standard", "月", cfg, "目录月价", "元/月/GiB")
        _insert_price(cur, product_code, hourly_price, "hourly", "小时", cfg, "目录小时价", "元/小时/GiB")
        if monthly_price > 0:
            prices += 1
        if hourly_price > 0:
            prices += 1
    return {"products": products, "prices": prices, "specs": specs, "skipped": skipped}


# ── Bandwidth importer ────────────────────────────────────────────────

def _import_bandwidth_sheet(ws, cfg: dict, cur) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    products = prices = specs = skipped = 0
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        if not any(vals):
            continue
        if idx == 0 or "公网" in vals[0] or "精品带宽" in vals[0]:
            continue
        name = vals[0]
        if not name:
            skipped += 1
            continue
        product_code = f"WY-{slug(cfg['category_code'])}-{slug(name)}"
        _insert_product(cur, product_code, name, cfg, name, model=name, unit="项", ext_json={})
        products += 1
        _insert_inventory(cur, product_code)
        _insert_chunk(cur, product_code, name, name, "无影,带宽规格")
        # Try read prices: usually at col 4/5
        for i in range(4, len(vals)):
            val = to_float(vals[i])
            if val > 0:
                _insert_price(cur, product_code, val, "standard", "月", cfg, "单价", "元")
                prices += 1
                break
    return {"products": products, "prices": prices, "specs": specs, "skipped": skipped}


# ── Traffic package importer ──────────────────────────────────────────

def _import_traffic_sheet(ws, cfg: dict, cur) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    products = prices = specs = skipped = 0
    period = "1个月"
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        if not any(vals):
            continue
        if "有效期" in vals[0]:
            period = vals[0].strip()
            continue
        if "流量包规格" in vals[0] or "规格" in vals[0]:
            continue
        if "价格" in vals[0]:
            continue
        name = vals[0]
        if not name:
            skipped += 1
            continue
        product_code = f"WY-{slug(cfg['category_code'])}-{slug(name)}-{slug(period)}"
        _insert_product(cur, product_code, name, cfg, f"{name} ({period})", model=name, unit="个", ext_json={"validity_period": period})
        products += 1
        _insert_inventory(cur, product_code)
        _insert_chunk(cur, product_code, name, f"{name} ({period})", "无影,流量包")
        for i in range(1, len(vals)):
            val = to_float(vals[i])
            if val > 0:
                _insert_price(cur, product_code, val, "package", period, cfg, "价格", "元")
                prices += 1
                break
    return {"products": products, "prices": prices, "specs": specs, "skipped": skipped}


# ── Addon / AD Connector / Cloud Drive / Core Hour (shared pattern) ──

def _import_addon_sheet(ws, cfg: dict, cur) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    products = prices = specs = skipped = 0
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        if not any(vals):
            continue
        if idx == 0 or "规格" in vals[0]:
            continue
        name = vals[0]
        if not name:
            skipped += 1
            continue
        hourly = to_float(vals[1] if len(vals) > 1 else "")
        product_code = f"WY-{slug(cfg['category_code'])}-{slug(name)}"
        _insert_product(cur, product_code, name, cfg, name, model=name, unit="个", ext_json={})
        products += 1
        _insert_inventory(cur, product_code)
        _insert_chunk(cur, product_code, name, name, f"无影,{cfg['product_type_name']}")
        _insert_price(cur, product_code, hourly, "hourly", "小时", cfg, "目录小时价", "元/小时")
        if hourly > 0:
            prices += 1
    return {"products": products, "prices": prices, "specs": specs, "skipped": skipped}


def _import_core_hour_sheet(ws, cfg: dict, cur) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    products = prices = specs = skipped = 0
    # Complex matrix; import as product codes with multiple price rows
    periods: list[str] = []
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        if not any(vals):
            continue
        if "核时包梯度" in vals[0]:
            continue
        if "有效期" in vals[0] or "" == vals[0] and any("月" in str(v) for v in vals):
            periods = [str(v).replace("月", "month") for v in vals[1:5] if v]
            continue
        name = vals[0]
        if not name or not name.isdigit():
            continue
        product_code = f"WY-{slug(cfg['category_code'])}-{name}"
        _insert_product(cur, product_code, f"核时包 {name}", cfg, f"核时包梯度 {name}", model=name, unit="个", ext_json={"gradient": name})
        products += 1
        _insert_inventory(cur, product_code)
        _insert_chunk(cur, product_code, f"核时包 {name}", f"核时包梯度 {name}", "无影,核时包")
        for i in range(1, len(vals)):
            val = to_float(vals[i])
            if val > 0 and (i - 1) < len(periods):
                _insert_price(cur, product_code, val, "package", periods[i - 1], cfg, periods[i - 1], "元")
                prices += 1
    return {"products": products, "prices": prices, "specs": specs, "skipped": skipped}


# ── Terminal importer ─────────────────────────────────────────────────

def _import_terminal_sheet(ws, cfg: dict, cur) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    products = prices = specs = skipped = 0
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        if not any(vals):
            continue
        if idx == 0 or "序号" in vals[0]:
            continue
        name = vals[1] if len(vals) > 1 else vals[0]
        param = vals[2] if len(vals) > 2 else ""
        list_price = to_float(vals[3] if len(vals) > 3 else "")
        remark = vals[4] if len(vals) > 4 else ""
        if not name:
            skipped += 1
            continue
        product_code = f"WY-{slug(cfg['category_code'])}-{slug(name)}"
        _insert_product(cur, product_code, name, cfg, param, model=name, unit="台", ext_json={"remark": remark})
        products += 1
        _insert_inventory(cur, product_code)
        _insert_chunk(cur, product_code, name, param, "无影,终端外设")
        _insert_spec(cur, product_code, "参数", param, None, cfg["product_type_name"], 1)
        specs += 1
        _insert_price(cur, product_code, list_price, "standard", "台", cfg, "刊例价", "元/台", allow_zero=True)
        prices += 1
    return {"products": products, "prices": prices, "specs": specs, "skipped": skipped}


# ── AI assistant (贾维斯) importer ────────────────────────────────────

def _import_ai_sheet(ws, cfg: dict, cur) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    products = prices = specs = skipped = 0
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        if not any(vals):
            continue
        if "项目" in vals[0] or "套餐" in vals[0]:
            continue
        item = vals[0]
        detail = vals[1] if len(vals) > 1 else ""
        version = vals[2] if len(vals) > 2 else ""
        list_price = to_float(vals[3] if len(vals) > 3 else "")
        if not version:
            skipped += 1
            continue
        product_code = f"WY-{slug(cfg['category_code'])}-{slug(item)}-{slug(version)}"
        name = f"{item} - {version}"
        _insert_product(cur, product_code, name, cfg, detail, model=version, unit="个", ext_json={"item": item, "version": version})
        products += 1
        _insert_inventory(cur, product_code)
        _insert_chunk(cur, product_code, name, detail, "无影,AI助手")
        _insert_price(cur, product_code, list_price, "standard", "月", cfg, "价格", "元")
        if list_price > 0:
            prices += 1
    return {"products": products, "prices": prices, "specs": specs, "skipped": skipped}


if __name__ == "__main__":
    result = import_wuying_excel(EXCEL_PATH)
    import json
    print(json.dumps(result, ensure_ascii=False, indent=2))
