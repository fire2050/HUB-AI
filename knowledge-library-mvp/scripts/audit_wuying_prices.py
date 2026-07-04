from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
import subprocess
import sys
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    import openpyxl
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl is required: {exc}")

DEFAULT_EXCEL = Path("/home/xhb-szwl/.openclaw/workspace/Knowledge-Library-MVP/latest_wuying.xlsx")
DEFAULT_DB = Path("/home/xhb-szwl/.openclaw/workspace/Knowledge-Library-MVP/hubai_quotes.db")
DEFAULT_REPORT = Path("/home/xhb-szwl/.openclaw/workspace/Knowledge-Library-MVP/reports/wuying_price_audit_report.md")

# Same canonical sheets as the importer. The full source workbook may contain
# many extra reference sheets; only these sheets are currently imported into
# product/product_price and therefore can be compared 1:1.
CANONICAL_SHEETS = [
    "B2-磁盘规格list价格（必选项）",
    "B3-带宽规格list价格（可选项）",
    "B4-流量包list价格（可选项）",
    "B5-AD Connector的list价格（可选项）",
    "B6-企业网盘list价格（可选项）",
    "B7-核时包list价格（可选项）",
    "C0-终端描述汇总",
    "D1教育办公",
    "D21国内120小时",
    "D21国内不限时长-办公1小时休眠",
    "D22国内图形200小时",
    "D22国内不限时长(1小时休眠)",
    "D23国内不限时长",
    "E贾维斯",
    "D2-带宽规格list价格（可选项） ",
]
COMPUTE_SHEETS = {
    "D1教育办公",
    "D21国内120小时",
    "D21国内不限时长-办公1小时休眠",
    "D22国内图形200小时",
    "D22国内不限时长(1小时休眠)",
    "D23国内不限时长",
}
PRICE_LABEL_MAP = {
    "一个月": "1month",
    "一年": "1year",
    "二年": "2year",
    "三年": "3year",
    "四年": "4year",
    "五年": "5year",
    "六年": "6year",
}


def norm(v: Any) -> str:
    return str(v or "").strip()


def to_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("￥", "").replace("元", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else default


def money_equal(a: float, b: float) -> bool:
    return abs(float(a) - float(b)) < 0.005


def excel_compute_prices(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    prices: list[dict[str, Any]] = []
    for idx, row in enumerate(rows):
        vals = [norm(x) for x in row]
        if not any(k in " ".join(vals) for k in ["计算规格名称", "计算规格（必选项）"]):
            continue
        header_row = idx
        period_row = idx + 1
        header = [norm(x) for x in rows[header_row]]
        periods = [norm(x).replace("\n", "") for x in rows[period_row]] if period_row < len(rows) else []
        name_col = next((i for i, v in enumerate(header) if "名称" in v), 0)
        period_cols: list[tuple[int, str, str]] = []
        for i, p in enumerate(periods):
            if p in PRICE_LABEL_MAP:
                period_cols.append((i, PRICE_LABEL_MAP[p], p))
            elif "小时" in p and "目录" in p:
                period_cols.append((i, "hourly", "目录小时价"))
        for row_i, data_row in enumerate(rows[period_row + 1 :], start=period_row + 2):
            data = [norm(x) for x in data_row]
            if not any(data):
                continue
            name = data[name_col] if name_col < len(data) else ""
            if not name:
                continue
            if any(k in name for k in ["计算规格名称", "存储规格名称", "带宽规格名称"]):
                continue
            if any(k in name for k in ["国内region存储规格", "国内region精品带宽规格", "国际region存储规格", "国际region带宽规格"]):
                break
            for col, price_type, label in period_cols:
                value = to_float(data[col] if col < len(data) else "")
                if value > 0:
                    prices.append({"sheet": ws.title, "name": name, "price_type": price_type, "billing_period": label, "price": value, "row": row_i})
    return prices


def db_prices(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT p.product_name AS name, p.product_code, p.product_type, p.source_sheet AS product_source_sheet,
               pp.source_sheet AS sheet, pp.price_type, pp.billing_period, pp.unit_price AS price, pp.unit_label
        FROM product p
        JOIN product_price pp ON pp.product_code = p.product_code AND pp.status = 'active'
        WHERE p.status = 'active'
        ORDER BY pp.source_sheet, p.product_name, pp.price_type, pp.id
        """
    ).fetchall()
    return [dict(r) for r in rows]


def compare_compute(wb, conn: sqlite3.Connection) -> dict[str, Any]:
    excel_rows: list[dict[str, Any]] = []
    for sheet in COMPUTE_SHEETS:
        if sheet in wb.sheetnames:
            excel_rows.extend(excel_compute_prices(wb[sheet]))
    db_rows = [r for r in db_prices(conn) if r["sheet"] in COMPUTE_SHEETS]

    excel_map: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    for r in excel_rows:
        excel_map[(r["sheet"], r["name"], r["price_type"])].append(r["price"])
    db_map: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    for r in db_rows:
        db_map[(r["sheet"], r["name"], r["price_type"])].append(float(r["price"]))

    mismatches = []
    missing_in_db = []
    extra_in_db = []
    for key in sorted(set(excel_map) | set(db_map)):
        xs = sorted(excel_map.get(key, []))
        ds = sorted(db_map.get(key, []))
        if not xs:
            extra_in_db.append({"key": key, "db_prices": ds})
        elif not ds:
            missing_in_db.append({"key": key, "excel_prices": xs})
        elif len(xs) != len(ds) or any(not money_equal(a, b) for a, b in zip(xs, ds)):
            mismatches.append({"key": key, "excel_prices": xs, "db_prices": ds})

    return {
        "excel_price_rows": len(excel_rows),
        "db_price_rows": len(db_rows),
        "matched_keys": len(set(excel_map) & set(db_map)),
        "missing_in_db": missing_in_db,
        "extra_in_db": extra_in_db,
        "mismatches": mismatches,
        "excel_by_sheet": dict(Counter(r["sheet"] for r in excel_rows)),
        "db_by_sheet": dict(Counter(r["sheet"] for r in db_rows)),
    }


def workbook_summary(wb) -> dict[str, Any]:
    return {
        "sheet_count": len(wb.sheetnames),
        "canonical_present": [s for s in CANONICAL_SHEETS if s in wb.sheetnames],
        "canonical_missing": [s for s in CANONICAL_SHEETS if s not in wb.sheetnames],
        "extra_sheets": [s for s in wb.sheetnames if s not in CANONICAL_SHEETS],
    }


def db_summary(conn: sqlite3.Connection) -> dict[str, Any]:
    conn.row_factory = sqlite3.Row
    counts = {}
    for table in ["product_category", "product", "product_price", "inventory", "product_spec", "knowledge_chunk"]:
        try:
            counts[table] = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        except sqlite3.Error:
            counts[table] = None
    return {
        "counts": counts,
        "price_types": [dict(r) for r in conn.execute("SELECT price_type, COUNT(*) AS count, MIN(unit_price) AS min_price, MAX(unit_price) AS max_price FROM product_price GROUP BY price_type ORDER BY price_type")],
        "source_sheets": [dict(r) for r in conn.execute("SELECT COALESCE(source_sheet,'') AS sheet, COUNT(*) AS count FROM product_price GROUP BY sheet ORDER BY sheet")],
    }


def _table_rows(conn: sqlite3.Connection, sql: str) -> list[dict[str, Any]]:
    conn.row_factory = sqlite3.Row
    return [dict(r) for r in conn.execute(sql).fetchall()]


def _counter_rows(rows: list[dict[str, Any]], keys: list[str]) -> Counter:
    def clean(value: Any) -> Any:
        if isinstance(value, float):
            return round(value, 6)
        return value if value is not None else ""

    return Counter(tuple(clean(row.get(k)) for k in keys) for row in rows)


def compare_roundtrip_import(excel_path: Path, current_db: Path) -> dict[str, Any]:
    """Import the Excel into a temporary DB, then compare with current DB.

    This covers every sheet supported by the production importer, not just the
    compute sheets that can be parsed independently in this audit script.
    """
    project_root = Path(__file__).resolve().parents[1]
    with tempfile.TemporaryDirectory(prefix="wuying-price-audit-") as td:
        temp_db = Path(td) / "audit_import.db"
        env = os.environ.copy()
        env["FINANCE_WUKONG_DB"] = str(temp_db)
        init = subprocess.run(
            [sys.executable, "scripts/init_hubai_db.py"],
            cwd=project_root,
            env=env,
            text=True,
            capture_output=True,
            check=False,
        )
        if init.returncode != 0:
            return {"success": False, "stage": "init", "stderr": init.stderr, "stdout": init.stdout}
        imp = subprocess.run(
            [sys.executable, "scripts/import_wuying_domestic_compute.py", str(excel_path)],
            cwd=project_root,
            env=env,
            text=True,
            capture_output=True,
            check=False,
        )
        if imp.returncode != 0:
            return {"success": False, "stage": "import", "stderr": imp.stderr, "stdout": imp.stdout}

        cur_conn = sqlite3.connect(current_db)
        tmp_conn = sqlite3.connect(temp_db)
        product_sql = """
            SELECT product_code, product_name, category_code, product_type, product_type_name,
                   region_scope, source_sheet, brand, model, unit, product_config_description, status
            FROM product
            ORDER BY product_code
        """
        price_sql = """
            SELECT p.product_name, pp.product_code, pp.unit_price, pp.currency, pp.price_type,
                   COALESCE(pp.billing_period,'') AS billing_period,
                   COALESCE(pp.source_sheet,'') AS source_sheet,
                   COALESCE(pp.source_column,'') AS source_column,
                   COALESCE(pp.unit_label,'') AS unit_label,
                   pp.status
            FROM product_price pp
            JOIN product p ON p.product_code = pp.product_code
            ORDER BY pp.source_sheet, pp.product_code, pp.price_type, pp.billing_period, pp.unit_price
        """
        cur_products = _table_rows(cur_conn, product_sql)
        tmp_products = _table_rows(tmp_conn, product_sql)
        cur_prices = _table_rows(cur_conn, price_sql)
        tmp_prices = _table_rows(tmp_conn, price_sql)
        cur_conn.close(); tmp_conn.close()

        product_keys = ["product_code", "product_name", "category_code", "product_type", "product_type_name", "region_scope", "source_sheet", "brand", "model", "unit", "product_config_description", "status"]
        price_keys = ["product_name", "product_code", "unit_price", "currency", "price_type", "billing_period", "source_sheet", "source_column", "unit_label", "status"]
        cur_product_counter = _counter_rows(cur_products, product_keys)
        tmp_product_counter = _counter_rows(tmp_products, product_keys)
        cur_price_counter = _counter_rows(cur_prices, price_keys)
        tmp_price_counter = _counter_rows(tmp_prices, price_keys)
        product_missing = list((tmp_product_counter - cur_product_counter).elements())
        product_extra = list((cur_product_counter - tmp_product_counter).elements())
        price_missing = list((tmp_price_counter - cur_price_counter).elements())
        price_extra = list((cur_price_counter - tmp_price_counter).elements())
        return {
            "success": True,
            "current_product_rows": len(cur_products),
            "excel_import_product_rows": len(tmp_products),
            "current_price_rows": len(cur_prices),
            "excel_import_price_rows": len(tmp_prices),
            "product_missing_in_current": len(product_missing),
            "product_extra_in_current": len(product_extra),
            "price_missing_in_current": len(price_missing),
            "price_extra_in_current": len(price_extra),
            "product_missing_samples": product_missing[:20],
            "product_extra_samples": product_extra[:20],
            "price_missing_samples": price_missing[:20],
            "price_extra_samples": price_extra[:20],
        }


def write_report(path: Path, excel_path: Path, db_path: Path, summary: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    comp = summary["compute_compare"]
    status = "通过" if not (comp["missing_in_db"] or comp["extra_in_db"] or comp["mismatches"]) else "存在差异"
    lines = [
        "# 无影产品 Excel 与数据库价格核对报告",
        "",
        f"- Excel：`{excel_path}`",
        f"- 数据库：`{db_path}`",
        f"- 核对结论（计算规格价格）：**{status}**",
        "",
        "## 数据库概览",
        "",
        "| 表 | 行数 |",
        "|---|---:|",
    ]
    for k, v in summary["db"]["counts"].items():
        lines.append(f"| {k} | {v} |")
    lines.extend(["", "## 价格类型分布", "", "| price_type | 数量 | 最小价 | 最大价 |", "|---|---:|---:|---:|"])
    for r in summary["db"]["price_types"]:
        lines.append(f"| {r['price_type']} | {r['count']} | {r['min_price']} | {r['max_price']} |")
    lines.extend(["", "## 计算规格逐价核对", "", "| 指标 | 数量 |", "|---|---:|", f"| Excel 价格行 | {comp['excel_price_rows']} |", f"| DB 价格行 | {comp['db_price_rows']} |", f"| 匹配 Key | {comp['matched_keys']} |", f"| DB 缺失 | {len(comp['missing_in_db'])} |", f"| DB 额外 | {len(comp['extra_in_db'])} |", f"| 价格不一致 | {len(comp['mismatches'])} |"])
    rt = summary.get("roundtrip_import", {})
    if rt:
        rt_status = "通过" if rt.get("success") and not any(rt.get(k, 0) for k in ["product_missing_in_current", "product_extra_in_current", "price_missing_in_current", "price_extra_in_current"]) else "存在差异"
        lines.extend([
            "",
            "## 全量导入结果核对（当前导入器覆盖的所有产品）",
            "",
            f"- 结论：**{rt_status}**",
            "",
            "| 指标 | 数量 |",
            "|---|---:|",
            f"| 当前 DB 产品行 | {rt.get('current_product_rows')} |",
            f"| Excel 临时导入产品行 | {rt.get('excel_import_product_rows')} |",
            f"| 当前 DB 价格行 | {rt.get('current_price_rows')} |",
            f"| Excel 临时导入价格行 | {rt.get('excel_import_price_rows')} |",
            f"| 当前 DB 缺失产品 | {rt.get('product_missing_in_current')} |",
            f"| 当前 DB 额外产品 | {rt.get('product_extra_in_current')} |",
            f"| 当前 DB 缺失价格 | {rt.get('price_missing_in_current')} |",
            f"| 当前 DB 额外价格 | {rt.get('price_extra_in_current')} |",
        ])
    lines.extend(["", "## Excel/DB Sheet 价格行分布", "", "| Sheet | Excel | DB |", "|---|---:|---:|"])
    for sheet in sorted(set(comp["excel_by_sheet"]) | set(comp["db_by_sheet"])):
        lines.append(f"| {sheet} | {comp['excel_by_sheet'].get(sheet,0)} | {comp['db_by_sheet'].get(sheet,0)} |")
    if comp["missing_in_db"] or comp["extra_in_db"] or comp["mismatches"]:
        lines.extend(["", "## 差异明细（前 50 条）", ""])
        for title, items in [("DB 缺失", comp["missing_in_db"]), ("DB 额外", comp["extra_in_db"]), ("价格不一致", comp["mismatches"] )]:
            if not items:
                continue
            lines.append(f"### {title}")
            for item in items[:50]:
                lines.append(f"- `{item}`")
            if len(items) > 50:
                lines.append(f"- ……共 {len(items)} 条")
    lines.extend(["", "## 工作簿 Sheet", "", f"- Sheet 总数：{summary['workbook']['sheet_count']}", f"- 已纳入当前导入器的 Sheet：{len(summary['workbook']['canonical_present'])}", f"- 当前未纳入导入器的额外 Sheet：{len(summary['workbook']['extra_sheets'])}"])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit Wuying Excel prices against HubAI DB")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    wb = openpyxl.load_workbook(args.excel, data_only=True, read_only=True)
    conn = sqlite3.connect(args.db)
    summary = {
        "workbook": workbook_summary(wb),
        "db": db_summary(conn),
        "compute_compare": compare_compute(wb, conn),
        "roundtrip_import": compare_roundtrip_import(args.excel, args.db),
    }
    write_report(args.report, args.excel, args.db, summary)
    conn.close()
    if args.json:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        comp = summary["compute_compare"]
        print({
            "report": str(args.report),
            "excel_price_rows": comp["excel_price_rows"],
            "db_price_rows": comp["db_price_rows"],
            "missing_in_db": len(comp["missing_in_db"]),
            "extra_in_db": len(comp["extra_in_db"]),
            "mismatches": len(comp["mismatches"]),
        })


if __name__ == "__main__":
    main()
